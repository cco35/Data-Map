"""
Tableau Lineage Mapper v5
=========================
Parses .twb/.twbx files — no Tableau Server required.

Five tabs:
  1. Tree View        — Collapsible workbook → worksheet → datasource hierarchy
  2. Table View       — Searchable/filterable flat table
  3. Impact Analysis  — Search a datasource/table, see every worksheet affected
  4. Cleanup          — No-source worksheets + duplicate datasource usage
  5. Diagrams         — Left-to-right family tree per workbook (Green→Blue→Coral)

Outputs:
  - tableau_lineage.xlsx  — Formatted Excel workbook
  - tableau_lineage.html  — Interactive HTML report

Usage:
    python tableau_lineage_mapper.py --folder /path/to/twb/files
    python tableau_lineage_mapper.py --folder . --output my_report

Requirements:
    pip install openpyxl
"""

import sys, json, zipfile, argparse, re
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────
#  NAMESPACE STRIPPING
# ─────────────────────────────────────────────

def _strip_namespaces(content: bytes) -> bytes:
    """
    Remove XML namespace declarations and prefixes from raw .twb bytes
    so that ElementTree xpath queries work with plain unprefixed tag names.
    Handles both single- and double-quoted xmlns attributes.
    """
    # Remove xmlns="..." double-quoted
    content = re.sub(rb'\s+xmlns(?::\w+)?="[^"]*"', b'', content)
    # Remove xmlns='...' single-quoted
    content = re.sub(rb"\s+xmlns(?::\w+)?='[^']*'", b'', content)
    # Strip prefixes from tags: <ns:tag> -> <tag>, </ns:tag> -> </tag>
    content = re.sub(rb'<(/?)[\w][\w.-]*:([\w][\w.-]*)', rb'<\1\2', content)
    # Strip prefixes from attributes: ns:attr="v" -> attr="v"
    content = re.sub(rb' [\w][\w.-]*:([\w][\w.-]*)=', rb' \1=', content)
    return content


# ─────────────────────────────────────────────
#  PARSING
# ─────────────────────────────────────────────

def get_twb_tree(filepath: Path):
    if filepath.suffix.lower() == ".twbx":
        with zipfile.ZipFile(filepath, "r") as zf:
            inner = [n for n in zf.namelist() if n.endswith(".twb")]
            if not inner:
                raise ValueError(f"No .twb found inside {filepath.name}")
            with zf.open(inner[0]) as f:
                content = _strip_namespaces(f.read())
        return ET.ElementTree(ET.fromstring(content))
    content = _strip_namespaces(filepath.read_bytes())
    return ET.ElementTree(ET.fromstring(content))


def extract_datasource_info(ds_element):
    info = {
        "ds_name":    ds_element.get("name", ""),
        "ds_caption": ds_element.get("caption", ""),
        "db_type": "", "server": "", "database": "", "schema": "",
        "table": "", "custom_sql": "", "is_view": False,
    }
    conn = ds_element.find(".//connection")
    if conn is not None:
        info["db_type"]  = conn.get("class", conn.get("dbclass", "")).lower()
        info["server"]   = conn.get("server", "")
        info["database"] = conn.get("dbname", conn.get("database", ""))
        info["schema"]   = conn.get("schema", "")

    rel_table = ds_element.find(".//relation[@type='table']")
    if rel_table is not None:
        tbl = rel_table.get("table", "").strip("[]")
        if tbl:
            info["table"] = tbl
            name_upper = tbl.upper()
            if (rel_table.get("view", "") == "true"
                    or name_upper.endswith("_VIEW")
                    or name_upper.endswith("_VW")
                    or name_upper.startswith("V_")
                    or name_upper.startswith("VW_")):
                info["is_view"] = True

    rel_sql = ds_element.find(".//relation[@type='text']")
    if rel_sql is not None and rel_sql.text:
        info["table"]      = "[Custom SQL]"
        info["custom_sql"] = rel_sql.text.strip()

    if not info["table"]:
        for rel in ds_element.findall(".//relation"):
            tbl = rel.get("table", "")
            if tbl:
                info["table"] = tbl.strip("[]")
                break

    if not info["db_type"] or not info["database"]:
        for nc in ds_element.findall(".//named-connections/named-connection"):
            c = nc.find("connection")
            if c is not None:
                cls = c.get("class", c.get("dbclass", "")).lower()
                if cls and cls != "federated":
                    if not info["db_type"]:
                        info["db_type"] = cls
                    if not info["database"]:
                        info["database"] = c.get("dbname", c.get("database", ""))
                    if not info["schema"]:
                        info["schema"] = c.get("schema", "")
                    if not info["server"]:
                        info["server"] = c.get("server", "")
                    break

    return info


def parse_workbook(filepath: Path):
    root = get_twb_tree(filepath).getroot()
    workbook_name = filepath.stem
    rows = []

    datasources = {}
    for ds in root.findall("datasources/datasource"):
        name = ds.get("name", "")
        if name.lower() in ("parameters", ""):
            continue
        datasources[name] = extract_datasource_info(ds)

    for ws in root.findall("worksheets/worksheet"):
        ws_name = ws.get("name", "Unknown Sheet")
        deps = set()
        for dep in ws.findall(".//datasource-dependencies"):
            ref = dep.get("datasource", "")
            if ref and ref.lower() != "parameters":
                deps.add(ref)
        for dep in ws.findall(".//view/datasources/datasource"):
            ref = dep.get("name", "")
            if ref and ref.lower() != "parameters":
                deps.add(ref)

        if not deps:
            rows.append({
                "workbook": workbook_name, "worksheet": ws_name,
                "ds_label": "(no datasource found)",
                "db_type": "", "server": "", "database": "",
                "schema": "", "table": "", "custom_sql": "", "is_view": False,
            })
        else:
            for ref in sorted(deps):
                ds = datasources.get(ref, {})
                label = ds.get("ds_caption") or ds.get("ds_name") or ref
                rows.append({
                    "workbook":   workbook_name, "worksheet":  ws_name,
                    "ds_label":   label,
                    "db_type":    ds.get("db_type", ""),
                    "server":     ds.get("server", ""),
                    "database":   ds.get("database", ""),
                    "schema":     ds.get("schema", ""),
                    "table":      ds.get("table", ""),
                    "custom_sql": ds.get("custom_sql", ""),
                    "is_view":    ds.get("is_view", False),
                })
    return rows


# ─────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────

_GREEN_DARK  = "0D2B1A"
_HEADER_BG   = "1A3D2B"
_HEADER_FG   = "6EE7B7"
_CORAL_DARK  = "2A1220"
_CORAL_MID   = "F87171"
_BLUE_DARK   = "0E1F35"
_BLUE_MID    = "4F8EF7"
_YELLOW_DARK = "2A2010"
_YELLOW_MID  = "FBBF24"
_GREY_DARK   = "1A1D27"
_GREY_MID    = "6B7280"
_ROW_ALT     = "F0FDF4"
_WHITE       = "FFFFFF"
_DARK_TEXT   = "1F2937"

def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _autofit_columns(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)

def _style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(name="Arial", bold=True, color=_HEADER_FG, size=10)
        cell.fill      = _fill(_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

def _object_type_style(obj_type):
    t = obj_type.lower()
    if "custom sql" in t: return _YELLOW_DARK, _YELLOW_MID
    if "view"       in t: return _CORAL_DARK,  _CORAL_MID
    if "table"      in t: return _BLUE_DARK,   _BLUE_MID
    return _GREY_DARK, _GREY_MID

def write_xlsx(rows, path: Path):
    if not HAS_OPENPYXL:
        print("  ⚠  openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    # ── Sheet 1: Full Lineage ──
    ws1 = wb.active
    ws1.title = "Full Lineage"
    headers = ["Workbook", "Worksheet", "Datasource", "DB Type",
               "Database", "Schema", "Table / View", "Object Type", "Custom SQL"]
    ws1.append(headers)
    _style_header_row(ws1, len(headers))
    ws1.freeze_panes = "A2"

    for i, r in enumerate(rows, start=2):
        fp = ".".join(p for p in [r.get("database",""), r.get("schema",""), r.get("table","")] if p)
        table_val = fp or r.get("table","") or "—"
        obj_type  = ("Custom SQL" if r.get("custom_sql")
                     else "View"  if r.get("is_view")
                     else "Table" if r.get("table") and r.get("table") != "(no datasource found)"
                     else "—")
        row_data = [
            r.get("workbook",""), r.get("worksheet",""), r.get("ds_label",""),
            r.get("db_type","").upper() or "—",
            r.get("database","") or "—", r.get("schema","") or "—",
            table_val, obj_type,
            "Yes" if r.get("custom_sql") else "No",
        ]
        ws1.append(row_data)
        row_fill = _fill(_ROW_ALT) if i % 2 == 0 else _fill(_WHITE)
        for col, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=i, column=col)
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col == 8:
                bg, fg = _object_type_style(obj_type)
                cell.fill = _fill(bg)
                cell.font = Font(name="Arial", bold=True, color=fg, size=9)
            elif col == 1:
                cell.fill = row_fill
                cell.font = Font(name="Arial", bold=True, color=_DARK_TEXT, size=9)
            elif col == 9:
                cell.fill = _fill(_YELLOW_DARK) if val == "Yes" else row_fill
                cell.font = Font(name="Arial", color=_YELLOW_MID if val == "Yes" else _DARK_TEXT, size=9)
            else:
                cell.fill = row_fill
                cell.font = Font(name="Arial", color=_DARK_TEXT, size=9)
    _autofit_columns(ws1)

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")
    sum_headers = ["Workbook", "Total Worksheets", "Total Datasources",
                   "Unique Tables/Views", "Custom SQL Count", "No-Source Sheets"]
    ws2.append(sum_headers)
    _style_header_row(ws2, len(sum_headers))
    ws2.freeze_panes = "A2"

    agg = defaultdict(lambda: {"sheets": set(), "ds": set(), "tables": set(), "sql": 0, "orphans": 0})
    for r in rows:
        wb_name = r.get("workbook","")
        agg[wb_name]["sheets"].add(r.get("worksheet",""))
        if r.get("ds_label") != "(no datasource found)":
            agg[wb_name]["ds"].add(r.get("ds_label",""))
            fp = ".".join(p for p in [r.get("database",""), r.get("schema",""), r.get("table","")] if p)
            if fp: agg[wb_name]["tables"].add(fp)
            if r.get("custom_sql"): agg[wb_name]["sql"] += 1
        else:
            agg[wb_name]["orphans"] += 1

    for i, (wb_name, data) in enumerate(sorted(agg.items()), start=2):
        row_data = [wb_name, len(data["sheets"]), len(data["ds"]),
                    len(data["tables"]), data["sql"], data["orphans"]]
        ws2.append(row_data)
        row_fill = _fill(_ROW_ALT) if i % 2 == 0 else _fill(_WHITE)
        for col, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=col)
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            cell.fill      = row_fill
            cell.font      = Font(name="Arial", bold=(col == 1),
                                  color="F87171" if col == 6 and isinstance(val, int) and val > 0 else _DARK_TEXT,
                                  size=9)
    _autofit_columns(ws2)

    # ── Sheet 3: No-Source Sheets ──
    ws3 = wb.create_sheet("No-Source Sheets")
    ws3.append(["Workbook", "Worksheet"])
    _style_header_row(ws3, 2)
    ws3.freeze_panes = "A2"
    orphans = [r for r in rows if r.get("ds_label") == "(no datasource found)"]
    if orphans:
        for i, r in enumerate(orphans, start=2):
            ws3.append([r.get("workbook",""), r.get("worksheet","")])
            row_fill = _fill(_ROW_ALT) if i % 2 == 0 else _fill(_WHITE)
            for col in range(1, 3):
                cell = ws3.cell(row=i, column=col)
                cell.border    = _thin_border()
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill      = row_fill
                cell.font      = Font(name="Arial", bold=(col == 1), color=_DARK_TEXT, size=9)
    else:
        ws3.append(["✅ No worksheets without a datasource found.", ""])
    _autofit_columns(ws3)

    wb.save(path)
    print(f"  ✓ XLSX → {path}")


# ─────────────────────────────────────────────
#  HTML
# ─────────────────────────────────────────────

def build_html(rows, path: Path):

    tree_data = defaultdict(lambda: defaultdict(list))
    for r in rows:
        tree_data[r["workbook"]][r["worksheet"]].append(r)

    nodes = []
    for wb, sheets in sorted(tree_data.items()):
        ws_nodes, all_tables = [], set()
        for ws, sources in sorted(sheets.items()):
            src_nodes = []
            for s in sources:
                fp = ".".join(filter(None, [s["database"], s["schema"], s["table"]]))
                src_nodes.append({
                    "label":      s["ds_label"],
                    "table":      fp or s["table"] or "—",
                    "db_type":    s["db_type"],
                    "is_view":    s["is_view"],
                    "custom_sql": (s["custom_sql"][:120]+"…") if len(s["custom_sql"])>120 else s["custom_sql"],
                })
                if fp: all_tables.add(fp)
            ws_nodes.append({"name": ws, "sources": src_nodes})
        nodes.append({"workbook": wb, "sheets": ws_nodes,
                      "table_count": len(all_tables), "sheet_count": len(sheets)})

    table_rows = []
    for r in rows:
        fp = ".".join(filter(None, [r["database"], r["schema"], r["table"]]))
        table_rows.append({
            "workbook":   r["workbook"], "worksheet":  r["worksheet"],
            "datasource": r["ds_label"], "db_type":    r["db_type"].upper() or "—",
            "full_path":  fp or r["table"] or "—",
            "is_view":    "View" if r["is_view"] else ("Custom SQL" if r["custom_sql"] else "Table"),
            "custom_sql": "Yes" if r["custom_sql"] else "No",
        })

    impact_map = defaultdict(list)
    for r in rows:
        if r["ds_label"] == "(no datasource found)": continue
        fp = ".".join(filter(None, [r["database"], r["schema"], r["table"]]))
        impact_map[r["ds_label"]].append({
            "workbook": r["workbook"], "worksheet": r["worksheet"],
            "full_path": fp or r["table"] or "—",
            "db_type": r["db_type"].upper() or "—",
            "custom_sql": bool(r["custom_sql"]),
        })
    impact_list = []
    for ds_label, affected in sorted(impact_map.items(), key=lambda x: -len(x[1])):
        seen, deduped = set(), []
        for a in affected:
            k = a["workbook"]+"|"+a["worksheet"]
            if k not in seen:
                seen.add(k); deduped.append(a)
        impact_list.append({
            "ds_label": ds_label, "affected": deduped,
            "ws_count": len(deduped),
            "wb_count": len(set(a["workbook"] for a in deduped)),
            "sample_path": deduped[0]["full_path"] if deduped else "—",
        })

    orphans = [{"workbook": r["workbook"], "worksheet": r["worksheet"]}
               for r in rows if r["ds_label"] == "(no datasource found)"]
    ds_table_map = defaultdict(list)
    for r in rows:
        if r["ds_label"] == "(no datasource found)": continue
        fp = ".".join(filter(None, [r["database"], r["schema"], r["table"]]))
        key = r["ds_label"]+" :: "+(fp or r["table"] or "—")
        ds_table_map[key].append({"workbook": r["workbook"], "worksheet": r["worksheet"]})
    duplicates = []
    for key, instances in sorted(ds_table_map.items()):
        seen, unique = set(), []
        for i in instances:
            k = i["workbook"]+"|"+i["worksheet"]
            if k not in seen:
                seen.add(k); unique.append(i)
        if len(unique) > 1:
            ds_label, fp = key.split(" :: ", 1)
            duplicates.append({"ds_label": ds_label, "path": fp,
                                "instances": unique, "count": len(unique)})
    duplicates.sort(key=lambda x: -x["count"])

    data_json   = json.dumps(nodes,       ensure_ascii=False)
    table_json  = json.dumps(table_rows,  ensure_ascii=False)
    impact_json = json.dumps(impact_list, ensure_ascii=False)
    orphan_json = json.dumps(orphans,     ensure_ascii=False)
    dupes_json  = json.dumps(duplicates,  ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Tableau Lineage Map</title>
<style>
:root{{
  --bg:#0f1117;--surface:#1a1d27;--surface2:#22263a;--border:#2e3348;
  --wb-stroke:#34d399;--wb-text:#6ee7b7;
  --ws-stroke:#4f8ef7;--ws-text:#93c5fd;
  --src-stroke:#f87171;--src-text:#fca5a5;
  --sql-stroke:#fbbf24;--sql-text:#fde68a;
  --none-stroke:#374151;--none-text:#6b7280;
  --green:#34d399;--accent:#4f8ef7;--yellow:#fbbf24;
  --red:#f87171;--orange:#fb923c;--coral:#fb7185;
  --text:#e2e8f0;--muted:#8892a4;
  --font:'Segoe UI',system-ui,sans-serif;--mono:'Cascadia Code','Fira Code',monospace;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--text);font-family:var(--font);font-size:14px}}
header{{background:linear-gradient(135deg,#111a14,#0f1117);border-bottom:1px solid var(--border);
  padding:18px 32px;display:flex;align-items:center;gap:14px}}
header h1{{font-size:20px;font-weight:700;letter-spacing:-.3px}}
header h1 span{{color:var(--green)}}
.pill{{background:var(--surface2);border:1px solid var(--border);border-radius:99px;
  padding:3px 12px;font-size:12px;color:var(--muted)}}
.tabs{{display:flex;border-bottom:1px solid var(--border);padding:0 32px;background:var(--surface);overflow-x:auto}}
.tab{{padding:12px 18px;cursor:pointer;font-size:13px;font-weight:500;color:var(--muted);
  border-bottom:2px solid transparent;transition:all .15s;user-select:none;white-space:nowrap}}
.tab:hover{{color:var(--text)}}.tab.active{{color:var(--green);border-bottom-color:var(--green)}}
main{{padding:24px 32px;max-width:1600px}}
.view{{display:none}}.view.active{{display:block}}
.stats{{display:flex;gap:12px;margin-bottom:24px;flex-wrap:wrap}}
.stat-card{{background:var(--surface);border:1px solid var(--border);border-radius:8px;
  padding:14px 20px;flex:1;min-width:130px}}
.stat-card .num{{font-size:26px;font-weight:700;color:var(--green)}}
.stat-card .lbl{{font-size:12px;color:var(--muted);margin-top:2px}}
.stat-card.warn .num{{color:var(--orange)}}.stat-card.danger .num{{color:var(--red)}}
.wb-card{{background:var(--surface);border:1px solid var(--wb-stroke);border-radius:10px;
  margin-bottom:10px;overflow:hidden;transition:box-shadow .2s}}
.wb-card:hover{{box-shadow:0 0 0 1px var(--wb-stroke)}}
.wb-hdr{{display:flex;align-items:center;gap:10px;padding:13px 18px;cursor:pointer;user-select:none}}
.wb-name{{font-weight:600;font-size:15px;flex:1;color:var(--wb-text)}}
.wb-meta{{color:var(--muted);font-size:12px;display:flex;gap:8px}}
.caret{{color:var(--muted);transition:transform .2s;font-size:12px;display:inline-block}}
.wb-body{{display:none;padding:0 18px 14px}}.wb-body.open{{display:block}}
.wb-hdr.open .caret{{transform:rotate(90deg)}}
.ws-row{{border-left:2px solid var(--ws-stroke);margin:5px 0 5px 8px;padding-left:14px}}
.ws-lbl{{display:flex;align-items:center;gap:8px;padding:5px 0;cursor:pointer;user-select:none}}
.ws-lbl:hover .ws-name{{color:var(--ws-text)}}
.ws-name{{font-size:13px;font-weight:500;color:var(--ws-text)}}
.ws-caret{{color:var(--muted);font-size:11px;transition:transform .2s;display:inline-block}}
.ws-lbl.open .ws-caret{{transform:rotate(90deg)}}
.ws-srcs{{display:none;padding-left:18px}}.ws-srcs.open{{display:block}}
.src-row{{display:grid;grid-template-columns:14px 1fr auto;gap:6px;align-items:start;
  padding:6px 0;border-top:1px solid var(--border)}}
.dot{{width:8px;height:8px;border-radius:50%;margin-top:4px;flex-shrink:0}}
.dot.table{{background:var(--coral)}}.dot.view{{background:var(--coral)}}
.dot.sql{{background:var(--yellow)}}.dot.none{{background:var(--muted)}}
.src-info{{font-size:12px}}.src-lbl{{font-weight:500;color:var(--src-text)}}
.src-path{{color:var(--accent);font-family:var(--mono);font-size:11px;margin-top:2px;word-break:break-all}}
.src-sql{{color:var(--yellow);font-size:11px;font-style:italic;margin-top:2px}}
.badge{{background:var(--surface2);border:1px solid var(--border);border-radius:4px;
  padding:1px 6px;font-size:11px;color:var(--accent);white-space:nowrap}}
.badge.view-badge{{color:var(--coral);border-color:var(--src-stroke)}}
.search-bar{{display:flex;gap:10px;margin-bottom:14px;align-items:center}}
.search-bar input,.search-bar select{{background:var(--surface);border:1px solid var(--border);
  border-radius:7px;padding:9px 14px;color:var(--text);font-size:13px;outline:none;transition:border-color .15s}}
.search-bar input{{flex:1}}.search-bar input::placeholder{{color:var(--muted)}}
.search-bar input:focus,.search-bar select:focus{{border-color:var(--green)}}
.row-count{{color:var(--muted);font-size:12px;margin-bottom:8px}}
table{{width:100%;border-collapse:collapse}}
thead tr{{background:var(--surface2)}}
th{{text-align:left;padding:10px 14px;font-size:12px;font-weight:600;color:var(--muted);
  text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid var(--border);white-space:nowrap}}
td{{padding:9px 14px;font-size:13px;border-bottom:1px solid var(--border);vertical-align:top}}
tr:hover td{{background:var(--surface2)}}
td.mono{{font-family:var(--mono);font-size:12px;color:var(--accent)}}
.chip{{display:inline-block;background:var(--surface2);border:1px solid var(--border);
  border-radius:4px;padding:1px 7px;font-size:11px}}
.chip.table{{color:var(--coral)}}.chip.view{{color:var(--coral)}}
.chip.sql{{color:var(--yellow)}}.chip.db{{color:#a78bfa}}
.impact-search{{margin-bottom:6px}}
.impact-search input{{width:100%;background:var(--surface);border:1px solid var(--border);
  border-radius:8px;padding:11px 16px;color:var(--text);font-size:14px;outline:none;
  transition:border-color .15s,box-shadow .15s}}
.impact-search input:focus{{border-color:var(--green);box-shadow:0 0 0 3px rgba(52,211,153,.12)}}
.impact-search input::placeholder{{color:var(--muted)}}
.impact-hint{{color:var(--muted);font-size:12px;margin-bottom:18px;margin-top:6px}}
.imp-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;margin-bottom:9px;overflow:hidden}}
.imp-hdr{{display:flex;align-items:center;gap:12px;padding:13px 18px;cursor:pointer;user-select:none;transition:background .15s}}
.imp-hdr:hover{{background:var(--surface2)}}
.imp-ds{{font-weight:600;font-size:14px;flex:1}}
.imp-path{{color:var(--muted);font-family:var(--mono);font-size:11px;margin-top:2px}}
.imp-body{{display:none;border-top:1px solid var(--border)}}.imp-body.open{{display:block}}
.imp-ws-row{{display:flex;align-items:center;gap:10px;padding:9px 18px;border-bottom:1px solid var(--border);font-size:13px}}
.imp-ws-row:last-child{{border-bottom:none}}
.imp-wb{{color:var(--muted);font-size:12px;min-width:200px;flex-shrink:0}}
.imp-badge{{border-radius:99px;padding:3px 10px;font-size:12px;font-weight:600;white-space:nowrap}}
.imp-badge.high{{background:rgba(248,113,113,.15);color:var(--red);border:1px solid rgba(248,113,113,.3)}}
.imp-badge.med{{background:rgba(251,191,36,.15);color:var(--yellow);border:1px solid rgba(251,191,36,.3)}}
.imp-badge.low{{background:rgba(52,211,153,.15);color:var(--green);border:1px solid rgba(52,211,153,.3)}}
.sql-tag{{background:rgba(251,191,36,.15);color:var(--yellow);border:1px solid rgba(251,191,36,.3);
  border-radius:4px;padding:1px 6px;font-size:11px;margin-left:4px}}
.section{{margin-bottom:36px}}
.section h2{{font-size:15px;font-weight:600;margin-bottom:4px;display:flex;align-items:center;gap:8px}}
.section p{{color:var(--muted);font-size:13px;margin-bottom:14px;line-height:1.6}}
.orphan-row{{display:flex;gap:12px;align-items:center;background:var(--surface);
  border:1px solid var(--border);border-radius:7px;padding:10px 14px;margin-bottom:6px}}
.orphan-wb{{color:var(--muted);font-size:12px;min-width:200px;flex-shrink:0}}
.dupe-card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;margin-bottom:9px;overflow:hidden}}
.dupe-hdr{{display:flex;align-items:center;gap:12px;padding:12px 16px;cursor:pointer;user-select:none;transition:background .15s}}
.dupe-hdr:hover{{background:var(--surface2)}}
.dupe-ds{{font-weight:600;font-size:13px;flex:1}}
.dupe-path{{font-family:var(--mono);font-size:11px;color:var(--accent);margin-top:2px}}
.dupe-body{{display:none;border-top:1px solid var(--border);padding:10px 16px}}.dupe-body.open{{display:block}}
.dupe-inst{{display:flex;gap:10px;padding:6px 0;font-size:13px;border-bottom:1px solid var(--border)}}
.dupe-inst:last-child{{border-bottom:none}}
.dupe-wb{{color:var(--muted);font-size:12px;min-width:200px;flex-shrink:0}}
.dupe-count{{background:rgba(251,113,133,.15);color:var(--coral);border:1px solid rgba(251,113,133,.3);
  border-radius:99px;padding:2px 10px;font-size:12px;font-weight:600;white-space:nowrap}}
.empty{{text-align:center;padding:40px;color:var(--muted);font-size:13px}}
.empty .icon{{font-size:30px;margin-bottom:8px}}
.diag-toolbar{{display:flex;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap}}
.diag-toolbar select{{background:var(--surface);border:1px solid var(--border);color:var(--text);
  border-radius:7px;padding:9px 14px;font-size:13px;outline:none;cursor:pointer;min-width:280px}}
.diag-toolbar select:focus{{border-color:var(--green)}}
.btn{{background:var(--green);color:#0a1f12;border:none;border-radius:7px;
  padding:9px 18px;font-size:13px;font-weight:700;cursor:pointer;transition:opacity .15s;white-space:nowrap}}
.btn:hover{{opacity:.85}}
.btn.secondary{{background:var(--surface2);border:1px solid var(--border);color:var(--text);font-weight:600}}
.diag-hint{{color:var(--muted);font-size:12px}}
.legend{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;
  background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 16px}}
.legend-item{{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--muted)}}
.legend-dot{{width:10px;height:10px;border-radius:3px;flex-shrink:0}}
.legend-dot.wb{{background:var(--wb-stroke)}}.legend-dot.ws{{background:var(--ws-stroke)}}
.legend-dot.src{{background:var(--src-stroke)}}.legend-dot.sql{{background:var(--yellow)}}
.legend-dot.none{{background:var(--none-stroke)}}
.diag-scroll{{overflow:auto;background:var(--surface);border:1px solid var(--border);
  border-radius:12px;padding:32px;min-height:300px}}
</style>
</head>
<body>
<header>
  <span style="font-size:24px">📊</span>
  <h1>Tableau <span>Lineage</span> Map</h1>
  <span class="pill" id="top-pill"></span>
</header>
<div class="tabs">
  <div class="tab active"  onclick="switchTab('tree')">🌲 Tree View</div>
  <div class="tab"         onclick="switchTab('table')">📋 Table View</div>
  <div class="tab"         onclick="switchTab('impact')">⚡ Impact Analysis</div>
  <div class="tab"         onclick="switchTab('cleanup')">🧹 Cleanup</div>
  <div class="tab"         onclick="switchTab('diagram')">🗺️ Diagrams</div>
</div>
<main>
  <div class="stats" id="stats"></div>
  <div class="view active" id="view-tree"><div id="tree-container"></div></div>
  <div class="view" id="view-table">
    <div class="search-bar">
      <input id="tbl-search" placeholder="Search workbook, worksheet, table…" oninput="filterTable()">
      <select id="tbl-wb" onchange="filterTable()"><option value="">All Workbooks</option></select>
    </div>
    <div class="row-count" id="tbl-count"></div>
    <table><thead><tr>
      <th>Workbook</th><th>Worksheet</th><th>Datasource</th>
      <th>DB Type</th><th>Full Path (DB.Schema.Table)</th><th>Object Type</th>
    </tr></thead><tbody id="tbl-body"></tbody></table>
  </div>
  <div class="view" id="view-impact">
    <div class="impact-search">
      <input id="imp-search" placeholder="Type a datasource name or table/view… e.g. ORDERS or SALES_DS" oninput="filterImpact()">
    </div>
    <div class="impact-hint">
      Each card = one datasource. Expand to see every worksheet that depends on it.
      Sorted by number of affected worksheets — highest risk first.
      🔴 10+ &nbsp;|&nbsp; 🟡 4–9 &nbsp;|&nbsp; 🟢 1–3
    </div>
    <div class="row-count" id="imp-count"></div>
    <div id="imp-container"></div>
  </div>
  <div class="view" id="view-cleanup">
    <div class="section">
      <h2>🔴 Worksheets With No Datasource <span id="orphan-badge"></span></h2>
      <p>No recognised data source. May be blank, broken, or an unsupported connection type. Good candidates for deletion.</p>
      <div id="orphan-container"></div>
    </div>
    <div class="section">
      <h2>🟡 Identical Datasource Across Multiple Worksheets <span id="dupe-badge"></span></h2>
      <p>Same datasource + table used in more than one worksheet. May be redundant — review for consolidation or removal.</p>
      <div id="dupe-container"></div>
    </div>
  </div>
  <div class="view" id="view-diagram">
    <div class="diag-toolbar">
      <select id="diag-select" onchange="drawDiagram()"><option value="">— Select a workbook —</option></select>
      <button class="btn secondary" onclick="prevWorkbook()">← Prev</button>
      <button class="btn secondary" onclick="nextWorkbook()">Next →</button>
      <button class="btn" onclick="saveSVG()">💾 Save as SVG</button>
      <span class="diag-hint" id="diag-hint"></span>
    </div>
    <div class="legend">
      <div class="legend-item"><div class="legend-dot wb"></div> Workbook</div>
      <div class="legend-item"><div class="legend-dot ws"></div> Worksheet</div>
      <div class="legend-item"><div class="legend-dot src"></div> Table / View</div>
      <div class="legend-item"><div class="legend-dot sql"></div> Custom SQL</div>
      <div class="legend-item"><div class="legend-dot none"></div> No datasource</div>
    </div>
    <div class="diag-scroll" id="diag-scroll">
      <div id="diag-placeholder" style="text-align:center;padding:60px;color:var(--muted)">
        <div style="font-size:40px;margin-bottom:12px">🗺️</div>
        Select a workbook above to view its lineage diagram
      </div>
      <svg id="diag-svg" style="display:none" xmlns="http://www.w3.org/2000/svg"></svg>
    </div>
  </div>
</main>
<script>
const TREE    = {data_json};
const ROWS    = {table_json};
const IMPACT  = {impact_json};
const ORPHANS = {orphan_json};
const DUPES   = {dupes_json};

const C = {{
  wb:   {{ fill:'#0d2b1a', stroke:'#34d399', text:'#6ee7b7' }},
  ws:   {{ fill:'#0e1f35', stroke:'#4f8ef7', text:'#93c5fd' }},
  src:  {{ fill:'#2a1220', stroke:'#f87171', text:'#fca5a5' }},
  sql:  {{ fill:'#2a2010', stroke:'#fbbf24', text:'#fde68a' }},
  none: {{ fill:'#1a1d27', stroke:'#374151', text:'#6b7280' }},
  edge: '#2e3348', bg: '#0f1117',
}};

function renderStats() {{
  const wbs    = new Set(ROWS.map(r=>r.workbook)).size;
  const sheets = new Set(ROWS.map(r=>r.workbook+'|'+r.worksheet)).size;
  const tables = new Set(ROWS.map(r=>r.full_path).filter(t=>t!=='—')).size;
  const sqls   = ROWS.filter(r=>r.custom_sql==='Yes').length;
  document.getElementById('top-pill').textContent = wbs+' workbooks';
  document.getElementById('stats').innerHTML = `
    <div class="stat-card"><div class="num">${{wbs}}</div><div class="lbl">Workbooks</div></div>
    <div class="stat-card"><div class="num">${{sheets}}</div><div class="lbl">Worksheets</div></div>
    <div class="stat-card"><div class="num">${{tables}}</div><div class="lbl">Unique Tables/Views</div></div>
    <div class="stat-card"><div class="num">${{sqls}}</div><div class="lbl">Custom SQL Sources</div></div>
    <div class="stat-card warn"><div class="num">${{DUPES.length}}</div><div class="lbl">Duplicate Sources</div></div>
    <div class="stat-card danger"><div class="num">${{ORPHANS.length}}</div><div class="lbl">No-Source Sheets</div></div>
  `;
}}

function renderTree() {{
  document.getElementById('tree-container').innerHTML = TREE.map((wb,wi) => `
    <div class="wb-card">
      <div class="wb-hdr" onclick="toggleEl('wbb-${{wi}}',this)">
        <span>📁</span><span class="wb-name">${{wb.workbook}}</span>
        <span class="wb-meta"><span>${{wb.sheet_count}} sheet${{wb.sheet_count!==1?'s':''}}</span><span>·</span><span>${{wb.table_count}} unique table${{wb.table_count!==1?'s':''}}</span></span>
        <span class="caret">▶</span>
      </div>
      <div class="wb-body" id="wbb-${{wi}}">
        ${{wb.sheets.map((ws,si) => `
          <div class="ws-row">
            <div class="ws-lbl" onclick="toggleEl('wss-${{wi}}-${{si}}',this)">
              <span class="ws-caret">▶</span><span class="ws-name">📄 ${{ws.name}}</span>
              <span style="color:var(--muted);font-size:11px;margin-left:6px">(${{ws.sources.length}} source${{ws.sources.length!==1?'s':''}})</span>
            </div>
            <div class="ws-srcs" id="wss-${{wi}}-${{si}}">
              ${{ws.sources.map(s=>{{
                const typ=s.custom_sql?'sql':(!s.table||s.table==='—')?'none':s.is_view?'view':'table';
                return `<div class="src-row">
                  <span class="dot ${{typ}}"></span>
                  <div class="src-info">
                    <div class="src-lbl">${{s.label}}</div>
                    <div class="src-path">${{s.table}}</div>
                    ${{s.custom_sql?`<div class="src-sql">⚡ Custom SQL: ${{s.custom_sql}}</div>`:''}}
                  </div>
                  <div style="display:flex;flex-direction:column;gap:3px;align-items:flex-end">
                    ${{s.db_type?`<span class="badge">${{s.db_type.toUpperCase()}}</span>`:''}}
                    ${{s.is_view?`<span class="badge view-badge">VIEW</span>`:''}}
                  </div>
                </div>`;
              }}).join('')}}
            </div>
          </div>`).join('')}}
      </div>
    </div>`).join('');
}}
function toggleEl(bodyId,hdrEl){{
  const body=document.getElementById(bodyId);
  const open=body.classList.toggle('open');
  if(hdrEl) hdrEl.classList.toggle('open',open);
}}

function renderTable(rows){{
  const tbody=document.getElementById('tbl-body');
  document.getElementById('tbl-count').textContent=rows.length+' row'+(rows.length!==1?'s':'');
  if(!rows.length){{tbody.innerHTML='<tr><td colspan="6" style="text-align:center;color:var(--muted);padding:32px">No results found</td></tr>';return;}}
  tbody.innerHTML=rows.map(r=>`<tr>
    <td><strong>${{r.workbook}}</strong></td><td>${{r.worksheet}}</td><td>${{r.datasource}}</td>
    <td><span class="chip db">${{r.db_type}}</span></td>
    <td class="mono">${{r.full_path}}</td>
    <td><span class="chip ${{r.is_view.toLowerCase()}}">${{r.is_view}}</span></td>
  </tr>`).join('');
}}
function filterTable(){{
  const q=document.getElementById('tbl-search').value.toLowerCase();
  const wb=document.getElementById('tbl-wb').value;
  renderTable(ROWS.filter(r=>(!wb||r.workbook===wb)&&(!q||Object.values(r).some(v=>String(v).toLowerCase().includes(q)))));
}}
function populateWBFilter(){{
  const sel=document.getElementById('tbl-wb');
  [...new Set(ROWS.map(r=>r.workbook))].sort().forEach(wb=>{{
    const o=document.createElement('option');o.value=wb;o.textContent=wb;sel.appendChild(o);
  }});
}}

function badgeClass(n){{return n>=10?'high':n>=4?'med':'low';}}
function renderImpact(items){{
  document.getElementById('imp-count').textContent=items.length+' datasource'+(items.length!==1?'s':'')+' found';
  const c=document.getElementById('imp-container');
  if(!items.length){{c.innerHTML='<div class="empty"><div class="icon">🔍</div>No matching datasources.</div>';return;}}
  c.innerHTML=items.map((item,i)=>`
    <div class="imp-card">
      <div class="imp-hdr" onclick="toggleImp(${{i}})">
        <span style="font-size:16px">🗄️</span>
        <div style="flex:1;min-width:0"><div class="imp-ds">${{item.ds_label}}</div>
          ${{item.sample_path!=='—'?`<div class="imp-path">${{item.sample_path}}</div>`:''}}
        </div>
        <div style="display:flex;gap:8px;align-items:center;flex-shrink:0">
          <span style="font-size:12px;color:var(--muted)">${{item.wb_count}} workbook${{item.wb_count!==1?'s':''}}</span>
          <span class="imp-badge ${{badgeClass(item.ws_count)}}">${{item.ws_count}} worksheet${{item.ws_count!==1?'s':''}} affected</span>
          <span class="caret" id="ic-${{i}}">▶</span>
        </div>
      </div>
      <div class="imp-body" id="icb-${{i}}">
        ${{item.affected.map(a=>`<div class="imp-ws-row">
          <span class="imp-wb">📁 ${{a.workbook}}</span><span>📄 ${{a.worksheet}}</span>
          ${{a.custom_sql?'<span class="sql-tag">Custom SQL</span>':''}}
        </div>`).join('')}}
      </div>
    </div>`).join('');
}}
function toggleImp(i){{
  const body=document.getElementById('icb-'+i);
  const open=body.classList.toggle('open');
  document.getElementById('ic-'+i).style.transform=open?'rotate(90deg)':'';
}}
function filterImpact(){{
  const q=document.getElementById('imp-search').value.toLowerCase().trim();
  renderImpact(!q?IMPACT:IMPACT.filter(item=>
    item.ds_label.toLowerCase().includes(q)||item.sample_path.toLowerCase().includes(q)||
    item.affected.some(a=>a.workbook.toLowerCase().includes(q)||a.worksheet.toLowerCase().includes(q))
  ));
}}

function renderCleanup(){{
  document.getElementById('orphan-badge').innerHTML=`<span class="chip src">${{ORPHANS.length}}</span>`;
  document.getElementById('orphan-container').innerHTML=!ORPHANS.length
    ?'<div class="empty"><div class="icon">✅</div>No worksheets without a datasource.</div>'
    :ORPHANS.map(o=>`<div class="orphan-row"><span class="orphan-wb">📁 ${{o.workbook}}</span><span>📄 ${{o.worksheet}}</span></div>`).join('');
  document.getElementById('dupe-badge').innerHTML=`<span class="chip sql">${{DUPES.length}}</span>`;
  document.getElementById('dupe-container').innerHTML=!DUPES.length
    ?'<div class="empty"><div class="icon">✅</div>No duplicate datasource usage found.</div>'
    :DUPES.map((d,i)=>`
      <div class="dupe-card">
        <div class="dupe-hdr" onclick="toggleDupe(${{i}})">
          <span>🔁</span>
          <div style="flex:1;min-width:0"><div class="dupe-ds">${{d.ds_label}}</div><div class="dupe-path">${{d.path}}</div></div>
          <span class="dupe-count">${{d.count}} worksheets</span>
          <span class="caret" id="dc-${{i}}">▶</span>
        </div>
        <div class="dupe-body" id="dcb-${{i}}">
          ${{d.instances.map(inst=>`<div class="dupe-inst"><span class="dupe-wb">📁 ${{inst.workbook}}</span><span>📄 ${{inst.worksheet}}</span></div>`).join('')}}
        </div>
      </div>`).join('');
}}
function toggleDupe(i){{
  const body=document.getElementById('dcb-'+i);
  const open=body.classList.toggle('open');
  document.getElementById('dc-'+i).style.transform=open?'rotate(90deg)':'';
}}

const TABS=['tree','table','impact','cleanup','diagram'];
function switchTab(name){{
  document.querySelectorAll('.tab').forEach((t,i)=>t.classList.toggle('active',TABS[i]===name));
  document.querySelectorAll('.view').forEach(v=>v.classList.remove('active'));
  document.getElementById('view-'+name).classList.add('active');
}}

// ── Diagram ──
const SVG_NS='http://www.w3.org/2000/svg';
const COL_W=[260,220,260];
const ROW_H=80;
const V_GAP=12;
const H_GAP=72;
const PAD=28;
const CHARS=26;

function svgEl(tag,attrs){{
  const el=document.createElementNS(SVG_NS,tag);
  for(const [k,v] of Object.entries(attrs)) el.setAttribute(k,v);
  return el;
}}
function svgLine(parent,x1,y1,x2,y2){{
  parent.appendChild(svgEl('path',{{
    d:`M ${{x1}} ${{y1}} C ${{(x1+x2)/2}} ${{y1}}, ${{(x1+x2)/2}} ${{y2}}, ${{x2}} ${{y2}}`,
    stroke:C.edge,'stroke-width':'1.5',fill:'none'
  }}));
}}
function wrapText(s,max){{
  if(!s) return ['—',null];
  if(s.length<=max) return [s,null];
  let cut=s.lastIndexOf(' ',max);
  if(cut<max*0.5) cut=max;
  const l1=s.slice(0,cut).trimEnd();
  let rest=s.slice(cut).trimStart();
  if(rest.length>max) rest=rest.slice(0,max-1)+'…';
  return [l1,rest||null];
}}
function nodeColors(src){{
  if(!src) return C.wb;
  if(src.custom_sql) return C.sql;
  if(!src.table||src.table==='—') return C.none;
  return C.src;
}}
function drawNode(svg,x,y,w,h,colors,rx,line1Full,line2,line3,fs1,fs2,tooltip){{
  const g=svgEl('g',{{}});
  g.appendChild(svgEl('rect',{{x,y,width:w,height:h,rx,fill:colors.fill,stroke:colors.stroke,'stroke-width':'1.5'}}));
  if(tooltip){{const t=document.createElementNS(SVG_NS,'title');t.textContent=tooltip;g.appendChild(t);}}
  const cx=x+12,midY=y+h/2;
  const [l1a,l1b]=wrapText(line1Full,CHARS);
  const total=1+(l1b?1:0)+(line2?1:0)+(line3?1:0);
  const lh=Math.min(16,(h-12)/total);
  let cy=midY-((total-1)*lh)/2;
  const mk=(txt,fs,fw,fill,ff)=>{{
    const t=svgEl('text',{{x:cx,y:cy,'dominant-baseline':'middle',
      'font-size':fs+'px','font-weight':fw,fill,
      'font-family':ff||'Segoe UI,sans-serif'}});
    t.textContent=txt;g.appendChild(t);cy+=lh;
  }};
  mk(l1a,fs1,'600',colors.text);
  if(l1b) mk(l1b,fs1,'600',colors.text);
  if(line2){{const [p1]=wrapText(line2,CHARS+4);mk(p1,fs2,'400','#4f8ef7','Cascadia Code,monospace');}}
  if(line3) mk(line3,9,'400',line3.includes('SQL')?'#fbbf24':line3.includes('VIEW')?'#fb7185':'#64748b','Cascadia Code,monospace');
  svg.appendChild(g);
}}
function drawWorkbookSVG(wb){{
  const svg=document.getElementById('diag-svg');
  svg.innerHTML='';
  const wsH=wb.sheets.map(ws=>Math.max(1,ws.sources.length)*(ROW_H+V_GAP)-V_GAP);
  const totalH=wsH.reduce((s,h)=>s+h,0)+(wb.sheets.length-1)*V_GAP;
  const canvasH=totalH+PAD*2;
  const canvasW=PAD+COL_W[0]+H_GAP+COL_W[1]+H_GAP+COL_W[2]+PAD;
  svg.setAttribute('width',canvasW);svg.setAttribute('height',canvasH);
  svg.setAttribute('viewBox',`0 0 ${{canvasW}} ${{canvasH}}`);svg.style.display='block';
  svg.appendChild(svgEl('rect',{{x:0,y:0,width:canvasW,height:canvasH,fill:C.bg}}));
  const colX=[PAD,PAD+COL_W[0]+H_GAP,PAD+COL_W[0]+H_GAP+COL_W[1]+H_GAP];
  const wbY=canvasH/2-ROW_H/2,wbCY=wbY+ROW_H/2,wbRX=colX[0]+COL_W[0];
  drawNode(svg,colX[0],wbY,COL_W[0],ROW_H,C.wb,12,
    '📁 '+wb.workbook,
    wb.sheet_count+' sheet'+(wb.sheet_count!==1?'s':'')+' · '+wb.table_count+' table'+(wb.table_count!==1?'s':''),
    null,13,11,wb.workbook);
  let curY=PAD;
  wb.sheets.forEach((ws,wi)=>{{
    const bh=wsH[wi],wsY=curY+bh/2-ROW_H/2,wsCY=wsY+ROW_H/2;
    const wsLX=colX[1],wsRX=colX[1]+COL_W[1];
    svgLine(svg,wbRX,wbCY,wsLX,wsCY);
    drawNode(svg,wsLX,wsY,COL_W[1],ROW_H,C.ws,8,
      '📄 '+ws.name,ws.sources.length+' source'+(ws.sources.length!==1?'s':''),
      null,12,10,ws.name);
    let srcY=curY;
    const srcs=ws.sources.length?ws.sources:[{{label:'(no datasource)',table:'',db_type:'',custom_sql:'',is_view:false}}];
    srcs.forEach(src=>{{
      const sCY=srcY+ROW_H/2,nc=nodeColors(src);
      const typeLabel=src.custom_sql?'⚡ Custom SQL':src.is_view?'◈ VIEW':src.db_type?src.db_type.toUpperCase():'';
      svgLine(svg,wsRX,wsCY,colX[2],sCY);
      const tip=[src.label,src.table,typeLabel].filter(Boolean).join('\\n');
      drawNode(svg,colX[2],srcY,COL_W[2],ROW_H,nc,6,
        src.label||'(no datasource)',src.table||'—',typeLabel,11,10,tip);
      srcY+=ROW_H+V_GAP;
    }});
    curY+=bh+V_GAP;
  }});
  document.getElementById('diag-placeholder').style.display='none';
  document.getElementById('diag-hint').textContent=
    wb.sheet_count+' worksheet'+(wb.sheet_count!==1?'s':'')+
    ' · '+wb.table_count+' unique table'+(wb.table_count!==1?'s':'');
}}
function drawDiagram(){{
  const sel=document.getElementById('diag-select');
  const idx=parseInt(sel.value);
  if(isNaN(idx)){{
    document.getElementById('diag-svg').style.display='none';
    document.getElementById('diag-placeholder').style.display='block';
    document.getElementById('diag-hint').textContent='';return;
  }}
  drawWorkbookSVG(TREE[idx]);
}}
function currentIdx(){{const v=document.getElementById('diag-select').value;return v===''?-1:parseInt(v);}}
function prevWorkbook(){{const i=currentIdx(),sel=document.getElementById('diag-select');
  if(i>0){{sel.value=i-1;drawDiagram();}}else if(i===-1&&TREE.length>0){{sel.value=TREE.length-1;drawDiagram();}}}}
function nextWorkbook(){{const i=currentIdx(),sel=document.getElementById('diag-select');
  if(i<TREE.length-1){{sel.value=i+1;drawDiagram();}}else if(i===-1&&TREE.length>0){{sel.value=0;drawDiagram();}}}}
function saveSVG(){{
  const svg=document.getElementById('diag-svg');
  if(svg.style.display==='none'){{alert('Please select a workbook first.');return;}}
  const wb=TREE[parseInt(document.getElementById('diag-select').value)];
  const name=(wb?wb.workbook:'diagram').replace(/[^a-z0-9_-]/gi,'_');
  const blob=new Blob(['<?xml version="1.0" encoding="UTF-8"?>'+svg.outerHTML],{{type:'image/svg+xml'}});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');a.href=url;a.download=name+'_lineage.svg';a.click();
  URL.revokeObjectURL(url);
}}
function populateDiagSelect(){{
  const sel=document.getElementById('diag-select');
  TREE.forEach((wb,i)=>{{
    const o=document.createElement('option');o.value=i;
    o.textContent=wb.workbook+' ('+wb.sheet_count+' sheets, '+wb.table_count+' tables)';
    sel.appendChild(o);
  }});
}}

renderStats();renderTree();populateWBFilter();renderTable(ROWS);
renderImpact(IMPACT);renderCleanup();populateDiagSelect();
</script>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ HTML → {path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Tableau workbook lineage mapper.")
    parser.add_argument("--folder", required=True, help="Folder with .twb/.twbx files")
    parser.add_argument("--output", default="tableau_lineage", help="Output filename prefix")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists():
        print(f"Error: '{folder}' does not exist."); sys.exit(1)

    files = sorted(list(folder.glob("*.twb")) + list(folder.glob("*.twbx")))
    if not files:
        print(f"No .twb or .twbx files found in '{folder}'."); sys.exit(1)

    print(f"\nFound {len(files)} workbook(s)\n")
    all_rows = []
    for f in files:
        print(f"  Parsing: {f.name}")
        try:
            rows = parse_workbook(f)
            all_rows.extend(rows)
            print(f"    → {len(rows)} links found")
        except Exception as e:
            print(f"    ✗ Error: {e}")

    if not all_rows:
        print("\nNo data extracted."); sys.exit(1)

    write_xlsx(all_rows, folder / f"{args.output}.xlsx")
    build_html(all_rows, folder / f"{args.output}.html")
    print(f"\nDone. Open {args.output}.html in any browser.")

if __name__ == "__main__":
    main()
